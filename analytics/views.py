import os
import traceback
from datetime import date, timedelta

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.core.management import call_command
from django.db import connection
from django.db.models import Sum, Count, Avg, Max, Min, F
from django.db.models.functions import TruncMonth

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken

from .models import (
    IndicateurPersonnalise, WidgetDashboard, ConfigurationProjet,
    DonneeBrute, EvenementComportemental,
)
from .kpi_engine import KPIEngine


# ============================================================
# VUES PRINCIPALES
# ============================================================

def index(request):
    return render(request, 'analytics/dashboard.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('index')
        return render(request, 'analytics/login.html', {'error': 'Identifiants invalides'})
    return render(request, 'analytics/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

def configurator(request):
    return render(request, 'analytics/configurator.html')

def dashboard(request):
    return render(request, 'analytics/dashboard.html')


# ============================================================
# API AUTHENTIFICATION
# ============================================================

@api_view(['POST'])
@permission_classes([AllowAny])
def api_login(request):
    try:
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            refresh = RefreshToken.for_user(user)
            return Response({'refresh': str(refresh), 'access': str(refresh.access_token)})
        return Response({'error': 'Identifiants invalides'}, status=401)
    except Exception as e:
        import traceback
        return Response({'error': f'Erreur serveur: {str(e)}', 'detail': traceback.format_exc()}, status=500)

def health_check(request):
    """Health check - plain Django view (no DRF, no DB, no auth)."""
    from datetime import datetime
    return HttpResponse(
        '{"status":"ok","timestamp":"' + datetime.now().isoformat() + '"}',
        content_type='application/json',
        status=200
    )

def db_check(request):
    """Diagnostic DB — vérifie la connexion et le compte admin."""
    import json as _json
    from datetime import datetime
    result = {'timestamp': datetime.now().isoformat(), 'db': 'unknown', 'admin_exists': False, 'error': None}
    try:
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
        result['db'] = 'ok'
        from django.contrib.auth import get_user_model
        User = get_user_model()
        result['admin_exists'] = User.objects.filter(username='admin').exists()
        result['user_count'] = User.objects.count()
    except Exception as e:
        result['db'] = 'error'
        result['error'] = str(e)
    return HttpResponse(_json.dumps(result), content_type='application/json', status=200)

@api_view(['POST'])
@permission_classes([AllowAny])
def api_reset_password(request):
    """Réinitialise le mot de passe d'un utilisateur (sans email)."""
    username = request.data.get('username', '').strip()
    new_password = request.data.get('new_password', '').strip()
    confirm_password = request.data.get('confirm_password', '').strip()

    if not username or not new_password or not confirm_password:
        return Response({'error': 'Tous les champs sont obligatoires.'}, status=400)

    if new_password != confirm_password:
        return Response({'error': 'Les mots de passe ne correspondent pas.'}, status=400)

    if len(new_password) < 6:
        return Response({'error': 'Le mot de passe doit contenir au moins 6 caractères.'}, status=400)

    User = get_user_model()
    try:
        user = User.objects.get(username=username)
    except User.DoesNotExist:
        return Response({'error': "Nom d'utilisateur introuvable."}, status=404)

    user.set_password(new_password)
    user.save()
    return Response({'success': 'Mot de passe réinitialisé avec succès.'}, status=200)


# ============================================================
# HELPERS
# ============================================================

def _get_ref_date():
    """Retourne la date la plus récente dans les données (évite de filtrer dans le vide si les données sont historiques)."""
    from django.db.models import Max
    result = DonneeBrute.objects.aggregate(max_date=Max('date_transaction'))
    return result['max_date'] or date.today()


def _apply_period(qs, periode, date_field='date_transaction', date_debut=None, date_fin=None):
    """Applique un filtre temporel. Utilise la date max des données comme référence (pas date.today())."""
    # Plage personnalisée
    if date_debut:
        try:
            qs = qs.filter(**{f'{date_field}__gte': date.fromisoformat(date_debut)})
        except (ValueError, TypeError):
            pass
    if date_fin:
        try:
            qs = qs.filter(**{f'{date_field}__lte': date.fromisoformat(date_fin)})
        except (ValueError, TypeError):
            pass
    if date_debut or date_fin:
        return qs

    # Périodes relatives : on ancre sur la date max des données, pas sur aujourd'hui
    if periode and periode != 'all':
        ref = _get_ref_date()
        if periode == 'mois':
            qs = qs.filter(**{f'{date_field}__year': ref.year, f'{date_field}__month': ref.month})
        elif periode == 'trimestre':
            q_start = ((ref.month - 1) // 3) * 3 + 1
            q_end = min(q_start + 2, 12)
            qs = qs.filter(**{
                f'{date_field}__year': ref.year,
                f'{date_field}__month__gte': q_start,
                f'{date_field}__month__lte': q_end,
            })
        elif periode == 'annee':
            qs = qs.filter(**{f'{date_field}__year': ref.year})
    return qs


def _base_qs(region=None, periode=None, config_id=None, date_debut=None, date_fin=None):
    qs = DonneeBrute.objects.all()
    if config_id:
        qs = qs.filter(config_id=config_id)
    if region and region != 'all':
        qs = qs.filter(region=region)
    qs = _apply_period(qs, periode, date_debut=date_debut, date_fin=date_fin)
    return qs


# ============================================================
# API KPIs DASHBOARD
# ============================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_kpis(request):
    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')

    qs = _base_qs(region=region, periode=periode, config_id=config_id, date_debut=date_debut, date_fin=date_fin)
    agg = qs.aggregate(ca=Sum('ca_ligne'), marge=Sum('marge_ligne'), nb=Count('id_donnee'))

    ca = float(agg['ca'] or 0)
    marge = float(agg['marge'] or 0)
    nb_cmd = agg['nb'] or 0
    nb_clients = qs.values('code_client').distinct().count()
    panier_moyen = round(ca / nb_cmd, 2) if nb_cmd else 0
    marge_pct = round((marge / ca * 100), 1) if ca else 0

    # Croissance vs mois précédent (ancré sur la date max des données)
    croissance = 0
    ref = _get_ref_date()
    prev_last = ref.replace(day=1) - timedelta(days=1)
    prev_first = prev_last.replace(day=1)
    qs_prev = DonneeBrute.objects.filter(date_transaction__gte=prev_first, date_transaction__lte=prev_last)
    if region and region != 'all':
        qs_prev = qs_prev.filter(region=region)
    if config_id:
        qs_prev = qs_prev.filter(config_id=config_id)
    ca_prev = float(qs_prev.aggregate(ca=Sum('ca_ligne'))['ca'] or 0)
    if ca_prev:
        croissance = round(((ca - ca_prev) / ca_prev) * 100, 1)

    return Response({
        'chiffre_affaires': round(ca, 2),
        'marge_totale': round(marge, 2),
        'marge_pourcentage': marge_pct,
        'nombre_commandes': nb_cmd,
        'panier_moyen': panier_moyen,
        'nombre_clients': nb_clients,
        'croissance': croissance,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_tendances(request):
    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')
    qs = _base_qs(region=region, periode=periode, config_id=config_id, date_debut=date_debut, date_fin=date_fin)

    data = (qs
        .annotate(mois=TruncMonth('date_transaction'))
        .values('mois')
        .annotate(ca=Sum('ca_ligne'), marge=Sum('marge_ligne'))
        .order_by('mois')
    )

    labels, ca_vals, marge_vals = [], [], []
    for row in data:
        if row['mois']:
            labels.append(row['mois'].strftime('%b %Y'))
            ca_vals.append(round(float(row['ca'] or 0), 2))
            marge_vals.append(round(float(row['marge'] or 0), 2))

    return Response({
        'labels': labels,
        'ca': ca_vals,
        'marge': marge_vals,
        'tendances': [{'mois': l, 'ca': c, 'marge': m} for l, c, m in zip(labels, ca_vals, marge_vals)],
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_par_region(request):
    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')
    qs = _base_qs(region=region, periode=periode, config_id=config_id, date_debut=date_debut, date_fin=date_fin)
    data = (qs
        .values('region')
        .annotate(ca=Sum('ca_ligne'), marge=Sum('marge_ligne'), nb_clients=Count('code_client', distinct=True))
        .order_by('-ca')
    )
    regions = [
        {'region': r['region'] or 'Non spécifié', 'ca': round(float(r['ca'] or 0), 2),
         'marge': round(float(r['marge'] or 0), 2), 'nb_clients': r['nb_clients']}
        for r in data
    ]
    return Response({'regions': regions})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_par_article(request):
    top = int(request.GET.get('top', 10))
    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')
    qs = _base_qs(region=region, periode=periode, config_id=config_id, date_debut=date_debut, date_fin=date_fin)
    data = (qs
        .values('code_article', 'nom_article', 'categorie')
        .annotate(ca=Sum('ca_ligne'), quantite=Sum('quantite'), nb_transactions=Count('id_donnee'))
        .order_by('-ca')[:top]
    )
    articles = [
        {'code': r['code_article'], 'nom': r['nom_article'] or r['code_article'],
         'categorie': r['categorie'] or '', 'ca': round(float(r['ca'] or 0), 2),
         'quantite': float(r['quantite'] or 0), 'nb_transactions': r['nb_transactions']}
        for r in data
    ]
    return Response({'articles': articles})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_funnel(request):
    # Utiliser EvenementComportemental si disponible
    types = [('vue_produit', 'Vue Produit'), ('ajout_panier', 'Ajout Panier'), ('achat', 'Achat')]
    funnel = []
    has_data = False
    for t, label in types:
        count = EvenementComportemental.objects.filter(type_evenement=t).count()
        if count:
            has_data = True
        funnel.append({'etape': label, 'count': count})

    if not has_data:
        # Proxy depuis DonneeBrute
        nb = DonneeBrute.objects.count()
        nb_clients = DonneeBrute.objects.values('code_client').distinct().count()
        funnel = [
            {'etape': 'Transactions', 'count': nb},
            {'etape': 'Clients uniques', 'count': nb_clients},
            {'etape': 'Articles uniques', 'count': DonneeBrute.objects.values('code_article').distinct().count()},
        ]

    return Response({'funnel': funnel})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_produits_fantomes(request):
    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')
    qs = _base_qs(region=region, periode=periode, config_id=config_id, date_debut=date_debut, date_fin=date_fin)
    # Articles avec beaucoup de transactions mais faible CA moyen par transaction
    data = (qs
        .values('code_article', 'nom_article')
        .annotate(nb_trans=Count('id_donnee'), ca_total=Sum('ca_ligne'), prix_moy=Avg('prix_unitaire'))
        .filter(nb_trans__gte=2)
        .order_by('ca_total')[:10]
    )
    fantomes = [
        {'code': r['code_article'], 'nom': r['nom_article'] or r['code_article'],
         'vues': r['nb_trans'] * 3, 'achats': r['nb_trans'],
         'ratio': round((r['nb_trans'] * 3) / max(r['nb_trans'], 1), 1),
         'prix': round(float(r['prix_moy'] or 0), 2)}
        for r in data
    ]
    return Response({'produits_fantomes': fantomes})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_produits_caches(request):
    """
    Produits 'cachés' : fort CA unitaire mais faible volume de transactions.
    Ces produits méritent plus de visibilité marketing.
    """
    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')
    qs = _base_qs(region=region, periode=periode, config_id=config_id, date_debut=date_debut, date_fin=date_fin)

    # Calculer les métriques par article
    data = list(qs.values('code_article', 'nom_article', 'categorie')
        .annotate(
            nb_trans=Count('id_donnee'),
            ca_total=Sum('ca_ligne'),
            prix_moy=Avg('prix_unitaire'),
            qte_total=Sum('quantite'),
        )
        .order_by('-prix_moy')
    )

    if not data:
        return Response({'produits_caches': []})

    # Calculer la médiane des transactions pour déterminer "faible volume"
    nb_trans_all = sorted([d['nb_trans'] for d in data])
    mediane = nb_trans_all[len(nb_trans_all) // 2]
    seuil = max(mediane, 2)

    # Produits cachés = prix élevé mais transactions sous la médiane
    caches = []
    for r in data:
        # Ratio d'opportunité : CA potentiel si on doublait les transactions
        ca_potentiel = float(r['ca_total'] or 0) * 2
        caches.append({
            'code': r['code_article'],
            'nom': r['nom_article'] or r['code_article'],
            'categorie': r['categorie'] or 'Autre',
            'nb_transactions': r['nb_trans'],
            'prix_moyen': round(float(r['prix_moy'] or 0), 2),
            'ca_actuel': round(float(r['ca_total'] or 0), 2),
            'ca_potentiel': round(ca_potentiel, 2),
            'score_opportunite': round(float(r['prix_moy'] or 0) / max(r['nb_trans'], 1) / 1000, 2),
            'statut': 'sous-exploite' if r['nb_trans'] <= seuil else 'normal',
        })

    # Trier par score d'opportunité décroissant
    caches.sort(key=lambda x: x['score_opportunite'], reverse=True)
    return Response({'produits_caches': caches[:10]})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_points_friction(request):
    """
    Points de friction estimés à partir des données transactionnelles.
    Identifie les articles avec forte variance de prix (indicateur de négociation/friction)
    et les clients avec longues périodes d'inactivité entre achats.
    """
    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')
    qs = _base_qs(region=region, periode=periode, config_id=config_id, date_debut=date_debut, date_fin=date_fin)

    if not qs.exists():
        return Response({'points_friction': [], 'source': 'aucune_donnee'})

    # Point de friction 1 : Articles avec forte dispersion de prix (renégociation fréquente)
    articles_stats = list(qs.values('nom_article', 'code_article')
        .annotate(
            prix_max=Max('prix_unitaire'),
            prix_min=Min('prix_unitaire'),
            nb_trans=Count('id_donnee'),
            ca=Sum('ca_ligne'),
        )
        .order_by('-nb_trans')
    )

    points = []
    for a in articles_stats:
        if a['prix_max'] and a['prix_min'] and float(a['prix_min']) > 0:
            dispersion = (float(a['prix_max']) - float(a['prix_min'])) / float(a['prix_max']) * 100
            if dispersion > 0:
                points.append({
                    'type': 'dispersion_prix',
                    'page': a['nom_article'] or a['code_article'],
                    'description': f"Variation de prix : {dispersion:.0f}% (min {a['prix_min']:,.0f} / max {a['prix_max']:,.0f} MAD)",
                    'abandons': a['nb_trans'],
                    'impact': round(float(a['ca'] or 0), 2),
                    'severite': 'haute' if dispersion > 20 else 'moyenne' if dispersion > 5 else 'faible',
                })

    # Point de friction 2 : Régions avec panier moyen faible vs total
    from django.db.models import FloatField
    ca_global = qs.aggregate(ca=Sum('ca_ligne'), nb=Count('id_donnee'))
    panier_global = float(ca_global['ca'] or 0) / max(ca_global['nb'] or 1, 1)

    regions = list(qs.values('region')
        .annotate(ca=Sum('ca_ligne'), nb=Count('id_donnee'))
        .order_by('region')
    )
    for r in regions:
        panier_region = float(r['ca'] or 0) / max(r['nb'] or 1, 1)
        if panier_region < panier_global * 0.8:
            points.append({
                'type': 'region_sous_performance',
                'page': f"Région {r['region']}",
                'description': f"Panier moyen {panier_region:,.0f} MAD vs moyenne {panier_global:,.0f} MAD ({(panier_region/panier_global-1)*100:.0f}%)",
                'abandons': r['nb'],
                'impact': round(float(r['ca'] or 0), 2),
                'severite': 'haute' if panier_region < panier_global * 0.6 else 'moyenne',
            })

    points.sort(key=lambda x: x['impact'], reverse=True)
    return Response({'points_friction': points[:10], 'source': 'donnees_transactionnelles'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_segmentation_comportementale(request):
    """
    Segmentation comportementale des clients basée sur les patterns d'achat.
    Utilise les données transactionnelles en l'absence de données comportementales web.
    """
    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')
    qs = _base_qs(region=region, periode=periode, config_id=config_id, date_debut=date_debut, date_fin=date_fin)

    if not qs.exists():
        return Response({'segments': [], 'clients': [], 'source': 'aucune_donnee'})

    # Date de référence = date max des données (pas aujourd'hui)
    ref_date = _get_ref_date()

    # Calculer les métriques par client
    clients_data = list(qs.values('code_client', 'nom_client')
        .annotate(
            nb_achats=Count('id_donnee'),
            ca_total=Sum('ca_ligne'),
            derniere_date=Max('date_transaction'),
            premiere_date=Min('date_transaction'),
            nb_articles=Count('code_article', distinct=True),
            nb_regions=Count('region', distinct=True),
        )
    )

    # Calcul des seuils percentiles une seule fois (hors de la boucle)
    ca_all_sorted = sorted(float(x.get('ca_total') or 0) for x in clients_data)
    nb_sorted = sorted(x.get('nb_achats') or 0 for x in clients_data)
    ca_p75 = ca_all_sorted[int(len(ca_all_sorted) * 0.75)] if ca_all_sorted else 0
    nb_p50 = nb_sorted[len(nb_sorted) // 2] if nb_sorted else 0

    # Recence médiane pour les seuils d'inactivité (relative aux données, pas à aujourd'hui)
    recence_all = []
    for c in clients_data:
        try:
            recence_all.append((ref_date - c['derniere_date']).days if c['derniere_date'] else 999)
        except Exception:
            recence_all.append(999)
    recence_all.sort()
    recence_p75 = recence_all[int(len(recence_all) * 0.75)] if recence_all else 180

    segments = {}
    clients = []

    for c in clients_data:
        try:
            recence_j = (ref_date - c['derniere_date']).days if c['derniere_date'] else 999
        except Exception:
            recence_j = 999

        nb_achats = c['nb_achats'] or 0
        ca_total  = float(c['ca_total'] or 0)
        nb_art    = c['nb_articles'] or 0

        # Segmentation avec seuils percentiles dynamiques
        if ca_total >= ca_p75 and nb_achats >= nb_p50:
            segment = 'grand_compte'
            label   = 'Grand compte'
            couleur = '#6366f1'
        elif nb_achats >= nb_p50 and nb_art >= 3:
            segment = 'acheteur_regulier'
            label   = 'Acheteur régulier'
            couleur = '#22c55e'
        elif recence_j > recence_p75 * 2:
            segment = 'client_inactif'
            label   = 'Client inactif'
            couleur = '#ef4444'
        elif recence_j > recence_p75:
            segment = 'risque_perte'
            label   = 'Risque de perte'
            couleur = '#f97316'
        elif nb_achats < nb_p50 and nb_art <= 2:
            segment = 'specialise'
            label   = 'Client spécialisé'
            couleur = '#06b6d4'
        elif nb_achats < nb_p50:
            segment = 'occasionnel'
            label   = 'Acheteur occasionnel'
            couleur = '#f59e0b'
        else:
            segment = 'standard'
            label   = 'Client standard'
            couleur = '#94a3b8'

        if segment not in segments:
            segments[segment] = {'segment': segment, 'label': label, 'couleur': couleur, 'count': 0, 'ca_total': 0}
        segments[segment]['count'] += 1
        segments[segment]['ca_total'] += ca_total

        clients.append({
            'code_client': c['code_client'],
            'nom_client': c['nom_client'] or c['code_client'],
            'segment': segment,
            'label_segment': label,
            'nb_achats': nb_achats,
            'ca_total': round(ca_total, 2),
            'nb_articles_distincts': nb_art,
            'recence_jours': recence_j,
        })

    seg_list = list(segments.values())
    for s in seg_list:
        s['ca_total'] = round(s['ca_total'], 2)

    clients.sort(key=lambda x: x['ca_total'], reverse=True)
    return Response({
        'segments': seg_list,
        'clients': clients,
        'total_clients': len(clients),
        'source': 'donnees_transactionnelles',
    })


# ============================================================
# API SEGMENTATION RFM
# ============================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_rfm(request):
    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')
    qs = _base_qs(region=region, periode=periode, config_id=config_id, date_debut=date_debut, date_fin=date_fin)

    # Ancrer la récence sur la date max des données, pas sur aujourd'hui
    ref_date = _get_ref_date()

    client_data = list(qs
        .values('code_client', 'nom_client')
        .annotate(
            derniere_transaction=Max('date_transaction'),
            frequence=Count('id_donnee'),
            montant=Sum('ca_ligne'),
        )
    )

    if not client_data:
        return Response({'segments': [], 'top_clients': [], 'total_clients': 0})

    # Calcul des percentiles pour des seuils adaptatifs RFM
    montants = sorted(float(c['montant'] or 0) for c in client_data)
    freqs = sorted(c['frequence'] or 0 for c in client_data)
    recences = []
    for c in client_data:
        try:
            recences.append((ref_date - c['derniere_transaction']).days if c['derniere_transaction'] else 999)
        except Exception:
            recences.append(999)
    recences.sort()

    n = len(client_data)
    def pct(lst, p): return lst[int(n * p)] if lst else 0

    m_p20, m_p40, m_p60, m_p80 = pct(montants, .2), pct(montants, .4), pct(montants, .6), pct(montants, .8)
    f_p20, f_p40, f_p60, f_p80 = pct(freqs, .2), pct(freqs, .4), pct(freqs, .6), pct(freqs, .8)
    r_p20, r_p40, r_p60, r_p80 = pct(recences, .2), pct(recences, .4), pct(recences, .6), pct(recences, .8)

    clients = []
    segment_counts = {}

    for c in client_data:
        try:
            recence = (ref_date - c['derniere_transaction']).days if c['derniere_transaction'] else 999
        except Exception:
            recence = 999
        frequence = c['frequence'] or 0
        montant = float(c['montant'] or 0)

        # Scores R, F, M sur percentiles (1-5) — adaptatifs aux données
        r = 5 if recence <= r_p20 else 4 if recence <= r_p40 else 3 if recence <= r_p60 else 2 if recence <= r_p80 else 1
        f = 5 if frequence >= f_p80 else 4 if frequence >= f_p60 else 3 if frequence >= f_p40 else 2 if frequence >= f_p20 else 1
        m = 5 if montant >= m_p80 else 4 if montant >= m_p60 else 3 if montant >= m_p40 else 2 if montant >= m_p20 else 1

        score = r + f + m

        if r >= 4 and f >= 4 and m >= 4:
            segment = 'champions'
        elif f >= 3 and m >= 3:
            segment = 'clients_fideles'
        elif r >= 3 and f <= 2:
            segment = 'clients_potentiels'
        elif r >= 4 and f <= 2:
            segment = 'nouveaux'
        elif r <= 2:
            segment = 'clients_perdus'
        else:
            segment = 'hibernation'

        segment_counts[segment] = segment_counts.get(segment, 0) + 1
        clients.append({
            'code_client': c['code_client'],
            'nom_client': c['nom_client'] or c['code_client'],
            'recence': recence,
            'frequence': frequence,
            'montant': round(montant, 2),
            'score_rfm': score,
            'segment_rfm': segment,
        })

    clients.sort(key=lambda x: x['score_rfm'], reverse=True)
    segments = [{'segment': k, 'count': v} for k, v in segment_counts.items()]

    return Response({'segments': segments, 'top_clients': clients[:20], 'total_clients': len(clients)})


# ============================================================
# API CHATBOT
# ============================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_chatbot(request):
    """
    Assistant analytique intelligent — Claude API avec contexte DB réel.
    Fallback automatique sur réponses règles si ANTHROPIC_API_KEY absent.
    """
    message_original = request.data.get('message', '').strip()
    message = message_original.lower()
    config_id = request.data.get('config_id')

    qs = DonneeBrute.objects.all()
    if config_id:
        qs = qs.filter(config_id=config_id)

    # ── Collecte du contexte réel depuis la base ──────────────────────────
    try:
        stats = qs.aggregate(
            ca=Sum('ca_ligne'), marge=Sum('marge_ligne'),
            nb=Count('id_donnee'), nb_clt=Count('code_client', distinct=True)
        )
        ca      = float(stats['ca']    or 0)
        marge   = float(stats['marge'] or 0)
        nb      = stats['nb']    or 0
        nb_clt  = stats['nb_clt'] or 0
        panier  = round(ca / nb, 0) if nb else 0
        taux_marge = round(marge / ca * 100, 1) if ca else 0

        top_regions = list(
            qs.values('region').annotate(ca=Sum('ca_ligne')).order_by('-ca')[:5]
        )
        top_articles = list(
            qs.values('nom_article', 'code_article')
              .annotate(ca=Sum('ca_ligne'), qte=Sum('quantite'))
              .order_by('-ca')[:5]
        )
        top_commerciaux = list(
            qs.values('nom_commercial', 'code_commercial')
              .annotate(ca=Sum('ca_ligne'))
              .order_by('-ca')[:5]
        )
        top_categories = list(
            qs.values('categorie').annotate(ca=Sum('ca_ligne')).order_by('-ca')[:5]
        )
        # Évolution mensuelle (6 derniers mois)
        evolution = list(
            qs.annotate(mois=TruncMonth('date_transaction'))
              .values('mois').annotate(ca=Sum('ca_ligne'))
              .order_by('-mois')[:6]
        )
        evolution.reverse()

    except Exception as e:
        return Response({'response': f"Erreur base de données : {e}", 'intent': 'error', 'confidence': 0})

    # ── Tentative d'appel Groq API (Llama 3.1 70B) ──────────────────────
    api_key = os.environ.get('GROQ_API_KEY', '')
    if api_key:
        try:
            from groq import Groq

            # Construction du contexte données
            regions_txt = '\n'.join(
                f"  • {r['region']}: {float(r['ca']):,.0f} MAD" for r in top_regions
            ) or "  Aucune donnée"
            articles_txt = '\n'.join(
                f"  • {a['nom_article'] or a['code_article']}: {float(a['ca']):,.0f} MAD ({int(a['qte'] or 0)} unités)"
                for a in top_articles
            ) or "  Aucune donnée"
            commerciaux_txt = '\n'.join(
                f"  • {c['nom_commercial'] or c['code_commercial']}: {float(c['ca']):,.0f} MAD"
                for c in top_commerciaux
            ) or "  Aucune donnée"
            categories_txt = '\n'.join(
                f"  • {c['categorie']}: {float(c['ca']):,.0f} MAD" for c in top_categories
            ) or "  Aucune donnée"
            evolution_txt = '\n'.join(
                f"  • {e['mois'].strftime('%B %Y') if e['mois'] else '?'}: {float(e['ca']):,.0f} MAD"
                for e in evolution
            ) or "  Aucune donnée"

            contexte = f"""=== DONNÉES RÉELLES DE LA BASE JUMIA ANALYTICS ===

INDICATEURS GLOBAUX :
  • Chiffre d'affaires total : {ca:,.0f} MAD
  • Marge totale : {marge:,.0f} MAD ({taux_marge}% du CA)
  • Nombre de transactions : {nb:,}
  • Clients actifs : {nb_clt:,}
  • Panier moyen : {panier:,.0f} MAD

TOP 5 RÉGIONS par CA :
{regions_txt}

TOP 5 PRODUITS par CA :
{articles_txt}

TOP 5 COMMERCIAUX par CA :
{commerciaux_txt}

TOP 5 CATÉGORIES par CA :
{categories_txt}

ÉVOLUTION MENSUELLE (6 derniers mois) :
{evolution_txt}
=== FIN DES DONNÉES ==="""

            client = Groq(api_key=api_key)
            resp = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                max_tokens=600,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "Tu es un assistant analytique expert intégré dans Jumia Analytics, "
                            "une plateforme d'analyse commerciale marocaine. "
                            "Tu as accès aux données réelles de l'entreprise fournies dans chaque message. "
                            "Réponds TOUJOURS en français, de façon concise, précise et professionnelle. "
                            "Utilise les chiffres exacts des données. Formule des observations pertinentes "
                            "et des recommandations actionnables quand c'est utile. "
                            "Ne dis jamais que tu n'as pas accès aux données."
                        )
                    },
                    {
                        "role": "user",
                        "content": f"{contexte}\n\nQuestion : {message_original}"
                    }
                ]
            )
            reponse = resp.choices[0].message.content
            return Response({
                'response': reponse,
                'intent': 'groq_ai',
                'confidence': 1.0,
                'message_original': message_original,
                'powered_by': 'llama-3.1-70b'
            })

        except Exception as e:
            # Erreur API → fallback règles (log pour debug)
            logger = __import__('logging').getLogger('analytics')
            logger.error(f"Groq API error: {e}")
            # Retourner l'erreur directement pour faciliter le debug
            return Response({
                'response': f"⚠️ Erreur assistant IA : {str(e)[:200]}. Contactez l'administrateur.",
                'intent': 'error',
                'confidence': 0,
                'powered_by': 'error'
            })

    # ── Fallback : réponses par règles (si pas de clé API) ───────────────
    def detecter_intention(msg):
        if any(w in msg for w in ['ca', "chiffre d'affaires", 'chiffre affaire', 'revenu', 'vente', 'total']):
            return 'ca_total'
        if any(w in msg for w in ['marge', 'profit', 'benefice', 'bénéfice']):
            return 'marge'
        if any(w in msg for w in ['client', 'acheteur']):
            return 'clients'
        if any(w in msg for w in ['region', 'région', 'zone', 'territoire']):
            return 'region'
        if any(w in msg for w in ['article', 'produit', 'top']):
            return 'article'
        if any(w in msg for w in ['commercial', 'vendeur', 'agent']):
            return 'commercial'
        if any(w in msg for w in ['panier', 'moyen']):
            return 'panier_moyen'
        if any(w in msg for w in ['bonjour', 'salut', 'hello', 'bonsoir']):
            return 'salutation'
        return 'fallback'

    intention = detecter_intention(message)
    if intention == 'salutation':
        reponse = "Bonjour ! Je suis l'assistant analytique Jumia Analytics. Posez-moi n'importe quelle question sur vos données commerciales !"
    elif intention == 'ca_total':
        reponse = f"Le chiffre d'affaires total est de {ca:,.0f} MAD sur {nb:,} transactions."
    elif intention == 'marge':
        reponse = f"La marge totale est de {marge:,.0f} MAD ({taux_marge}% du CA de {ca:,.0f} MAD)."
    elif intention == 'clients':
        reponse = f"Vous avez {nb_clt:,} client(s) actif(s), panier moyen : {panier:,.0f} MAD."
    elif intention == 'region':
        r = top_regions[0] if top_regions else None
        reponse = f"Meilleure région : {r['region']} ({float(r['ca']):,.0f} MAD)." if r else "Aucune donnée région."
    elif intention == 'article':
        a = top_articles[0] if top_articles else None
        reponse = f"Top produit : '{a['nom_article'] or a['code_article']}' ({float(a['ca']):,.0f} MAD)." if a else "Aucune donnée produit."
    elif intention == 'commercial':
        c = top_commerciaux[0] if top_commerciaux else None
        reponse = f"Meilleur commercial : '{c['nom_commercial'] or c['code_commercial']}' ({float(c['ca']):,.0f} MAD)." if c else "Aucune donnée."
    elif intention == 'panier_moyen':
        reponse = f"Panier moyen : {panier:,.0f} MAD ({nb:,} transactions, CA {ca:,.0f} MAD)."
    else:
        reponse = f"CA : {ca:,.0f} MAD | Marge : {taux_marge}% | Clients : {nb_clt:,} | Panier : {panier:,.0f} MAD. Posez une question plus précise !"
        intention = 'fallback'

    return Response({
        'response': reponse,
        'intent': intention,
        'confidence': 0.8,
        'message_original': message_original,
        'powered_by': 'rules'
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_chatbot_history(request):
    return Response({'history': []})


# ============================================================
# API PRÉVISIONS ML
# ============================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_previsions(request):
    """
    Prévisions CA sur 3 mois via régression linéaire + saisonnalité.
    Méthode hybride :
    - Tendance : régression linéaire sur l'historique complet
    - Saisonnalité : ratio mois N vs moyenne annuelle des années précédentes
    - Score qualité : R² ajusté avec interprétation lisible
    """
    region = request.GET.get('region', 'all')
    config_id = request.GET.get('config_id')
    # Les prévisions portent sur tout l'historique (pas de filtre période)
    qs = _base_qs(region=region, config_id=config_id)

    data = list(qs
        .annotate(mois=TruncMonth('date_transaction'))
        .values('mois')
        .annotate(ca=Sum('ca_ligne'))
        .order_by('mois')
    )

    labels  = [r['mois'].strftime('%b %Y') for r in data if r['mois']]
    ca_vals = [round(float(r['ca'] or 0), 2) for r in data if r['mois']]

    if len(ca_vals) < 3:
        return Response({
            'error': 'Données insuffisantes (minimum 3 mois requis)',
            'historique': {'labels': labels, 'ca': ca_vals},
            'previsions': [], 'r2_score': 0, 'qualite': 'insuffisant',
        })

    try:
        n = len(ca_vals)

        # ── 1. Régression linéaire (tendance) ──────────────────────────
        x_mean = (n - 1) / 2
        y_mean = sum(ca_vals) / n
        num = sum((i - x_mean) * (ca_vals[i] - y_mean) for i in range(n))
        den = sum((i - x_mean) ** 2 for i in range(n))
        slope     = num / den if den else 0
        intercept = y_mean - slope * x_mean

        # ── 2. Indices saisonniers ──────────────────────────────────────
        # Calculer la moyenne de CA par numéro de mois (1-12)
        mois_num = [r['mois'].month for r in data if r['mois']]
        ca_par_mois = {}
        count_par_mois = {}
        for i, m in enumerate(mois_num):
            ca_par_mois[m] = ca_par_mois.get(m, 0) + ca_vals[i]
            count_par_mois[m] = count_par_mois.get(m, 0) + 1

        moy_par_mois = {m: ca_par_mois[m] / count_par_mois[m] for m in ca_par_mois}
        moy_globale  = y_mean if y_mean else 1

        # Indice saisonnier = moyenne du mois / moyenne globale
        indices_saisonniers = {m: (v / moy_globale) for m, v in moy_par_mois.items()}

        # ── 3. Prévisions (tendance × saisonnalité) ────────────────────
        last_mois = data[-1]['mois']
        previsions = []
        for i in range(1, 4):
            y_pred, m_pred = last_mois.year, last_mois.month + i
            if m_pred > 12:
                m_pred -= 12
                y_pred += 1

            tendance  = slope * (n + i - 1) + intercept
            saisonnalite = indices_saisonniers.get(m_pred, 1.0)
            prediction = max(0, tendance * saisonnalite)

            next_label = date(y_pred, m_pred, 1).strftime('%b %Y')
            previsions.append({
                'mois': next_label,
                'prediction': round(prediction, 2),
                'tendance': round(tendance, 2),
                'indice_saisonnier': round(saisonnalite, 3),
            })

        # ── 4. Score R² sur les valeurs ajustées par saisonnalité ──────
        valeurs_ajustees = []
        for i, m in enumerate(mois_num):
            idx = indices_saisonniers.get(m, 1.0)
            val_aj = (slope * i + intercept) * idx
            valeurs_ajustees.append(val_aj)

        ss_res = sum((ca_vals[i] - valeurs_ajustees[i]) ** 2 for i in range(n))
        ss_tot = sum((ca_vals[i] - y_mean) ** 2 for i in range(n))
        r2 = round(max(0, 1 - ss_res / ss_tot), 3) if ss_tot else 0

        # Interprétation qualitative du R²
        if r2 >= 0.7:
            qualite = 'excellent'
        elif r2 >= 0.5:
            qualite = 'bon'
        elif r2 >= 0.3:
            qualite = 'modere'
        else:
            qualite = 'faible'

        # ── 5. Statistiques descriptives ───────────────────────────────
        ca_sorted = sorted(ca_vals)
        ca_median = ca_sorted[n // 2]
        croissance_moy = 0
        if n > 1:
            diffs = [(ca_vals[i] - ca_vals[i-1]) / max(ca_vals[i-1], 1) for i in range(1, n)]
            croissance_moy = round(sum(diffs) / len(diffs) * 100, 2)

        return Response({
            'historique': {'labels': labels, 'ca': ca_vals},
            'previsions': previsions,
            'r2_score': r2,
            'qualite': qualite,
            'methode': 'regression_saisonnalite',
            'stats': {
                'tendance_mensuelle': round(slope, 2),
                'ca_moyen': round(y_mean, 2),
                'ca_median': round(ca_median, 2),
                'croissance_mensuelle_moy': croissance_moy,
                'nb_mois_historique': n,
            },
        })
    except Exception as e:
        return Response({
            'error': str(e),
            'historique': {'labels': labels, 'ca': ca_vals},
            'previsions': [], 'r2_score': 0, 'qualite': 'erreur',
        })


# ============================================================
# API RECOMMANDATIONS DÉCISIONNELLES
# ============================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_recommandations(request):
    """
    Moteur de recommandations décisionnelles.
    Analyse l'ensemble des données et génère des préconisations
    actionnables classées par priorité et catégorie.
    """
    region    = request.GET.get('region', 'all')
    config_id = request.GET.get('config_id')
    qs = _base_qs(region=region, config_id=config_id)

    if not qs.exists():
        return Response({'recommandations': [], 'score_sante': 0, 'resume': 'Aucune donnée disponible.'})

    recos = []
    ref_date = _get_ref_date()

    # ── Agrégats globaux ──────────────────────────────────────
    from django.db.models import StdDev
    agg = qs.aggregate(
        ca=Sum('ca_ligne'), marge=Sum('marge_ligne'),
        nb=Count('id_donnee'), nb_clients=Count('code_client', distinct=True),
    )
    ca_total   = float(agg['ca']    or 0)
    marge_total= float(agg['marge'] or 0)
    nb_tx      = agg['nb'] or 1
    nb_clients = agg['nb_clients'] or 1
    marge_pct  = (marge_total / ca_total * 100) if ca_total else 0
    panier_moy = ca_total / nb_tx

    # ── Tendance mensuelle (6 derniers mois) ──────────────────
    monthly = list(qs
        .annotate(mois=TruncMonth('date_transaction'))
        .values('mois').annotate(ca=Sum('ca_ligne')).order_by('mois')
    )
    ca_vals = [float(m['ca'] or 0) for m in monthly]
    tendance_pct = 0
    if len(ca_vals) >= 2:
        last6 = ca_vals[-6:]
        if len(last6) >= 2 and last6[0]:
            tendance_pct = round((last6[-1] - last6[0]) / last6[0] * 100, 1)

    # ── Clients RFM ───────────────────────────────────────────
    clients_data = list(qs.values('code_client', 'nom_client')
        .annotate(
            nb_achats=Count('id_donnee'),
            ca_total=Sum('ca_ligne'),
            derniere_date=Max('date_transaction'),
        )
    )
    recences = []
    for c in clients_data:
        try:
            recences.append((ref_date - c['derniere_date']).days)
        except Exception:
            recences.append(999)
    rec_sorted = sorted(recences)
    rec_med = rec_sorted[len(rec_sorted)//2] if rec_sorted else 90

    clients_risque   = [c for c, r in zip(clients_data, recences) if r > rec_med * 1.5]
    clients_inactifs = [c for c, r in zip(clients_data, recences) if r > rec_med * 2.5]
    ca_risque = sum(float(c['ca_total'] or 0) for c in clients_risque)

    # ── Articles ──────────────────────────────────────────────
    articles = list(qs.values('code_article', 'nom_article', 'categorie')
        .annotate(ca=Sum('ca_ligne'), nb=Count('id_donnee'), prix_moy=Avg('prix_unitaire'))
        .order_by('-ca')
    )
    if articles:
        ca_top3   = sum(float(a['ca'] or 0) for a in articles[:3])
        conc_top3 = (ca_top3 / ca_total * 100) if ca_total else 0
        nb_all    = [a['nb'] for a in articles]
        nb_med    = sorted(nb_all)[len(nb_all)//2] if nb_all else 5
        articles_caches = [a for a in articles if a['nb'] <= nb_med and float(a['prix_moy'] or 0) > panier_moy]
    else:
        conc_top3 = 0
        articles_caches = []

    # ── Régions ───────────────────────────────────────────────
    regions_data = list(qs.values('region')
        .annotate(ca=Sum('ca_ligne'), nb=Count('id_donnee'))
        .order_by('-ca')
    )
    regions_faibles = []
    if regions_data:
        ca_moy_region = ca_total / len(regions_data)
        regions_faibles = [r for r in regions_data if float(r['ca'] or 0) < ca_moy_region * 0.6]

    # ═══════════════════════════════════════════════════
    # GÉNÉRATION DES RECOMMANDATIONS
    # ═══════════════════════════════════════════════════

    # 1. CLIENTS À RISQUE DE PERTE
    if clients_risque:
        ca_pct = round(ca_risque / ca_total * 100, 1) if ca_total else 0
        recos.append({
            'id': 'reco_clients_risque',
            'priorite': 'haute' if ca_pct > 20 else 'moyenne',
            'categorie': 'client',
            'icone': 'fa-user-clock',
            'couleur': '#ef4444',
            'titre': f'Réactiver {len(clients_risque)} client(s) à risque',
            'description': f'{len(clients_risque)} client(s) n\'ont pas commandé depuis plus de {int(rec_med * 1.5)} jours, représentant {ca_pct}% du CA total ({int(ca_risque/1000)}K MAD).',
            'actions': [
                f'Contacter en priorité : {", ".join(c["nom_client"] or c["code_client"] for c in clients_risque[:3])}',
                'Proposer une offre de réactivation (remise 10-15% sur prochaine commande)',
                'Organiser une visite commerciale dans les 2 semaines',
                'Analyser la raison de l\'inactivité (concurrent, prix, service ?)',
            ],
            'impact_estime': f'+{ca_pct:.0f}% CA potentiel récupérable',
            'donnees': {'nb_clients': len(clients_risque), 'ca_risque': round(ca_risque, 0), 'recence_seuil': int(rec_med * 1.5)},
        })

    # 2. CONCENTRATION ARTICLES
    if conc_top3 > 60:
        recos.append({
            'id': 'reco_diversification',
            'priorite': 'haute' if conc_top3 > 75 else 'moyenne',
            'categorie': 'commercial',
            'icone': 'fa-exclamation-triangle',
            'couleur': '#f97316',
            'titre': f'Risque de concentration : {conc_top3:.0f}% du CA sur 3 articles',
            'description': f'Votre activité dépend fortement de 3 articles qui génèrent {conc_top3:.0f}% du CA. Une rupture ou perte de contrat sur ces références fragilise l\'ensemble de l\'activité.',
            'actions': [
                f'Diversifier le portefeuille produit au-delà des top 3 articles',
                'Identifier 2-3 nouveaux articles à fort potentiel à promouvoir',
                'Négocier des contrats cadres pour sécuriser ces références clés',
                'Fixer un objectif de réduction de concentration à 50% sous 6 mois',
            ],
            'impact_estime': 'Réduction du risque commercial, CA plus résilient',
            'donnees': {'concentration_pct': round(conc_top3, 1), 'top3_ca': round(ca_top3, 0)},
        })

    # 3. ARTICLES SOUS-EXPLOITÉS
    if articles_caches:
        ca_pot = sum(float(a['ca'] or 0) * 1.5 for a in articles_caches[:3])
        recos.append({
            'id': 'reco_articles_caches',
            'priorite': 'moyenne',
            'categorie': 'marketing',
            'icone': 'fa-gem',
            'couleur': '#8b5cf6',
            'titre': f'{len(articles_caches)} article(s) à fort potentiel sous-exploités',
            'description': f'Ces articles ont un prix unitaire élevé mais un faible volume de commandes. Une meilleure visibilité commerciale pourrait générer {int(ca_pot/1000)}K MAD supplémentaires.',
            'actions': [
                f'Mettre en avant : {", ".join(a["nom_article"] or a["code_article"] for a in articles_caches[:3])}',
                'Inclure ces références dans les propositions commerciales systématiquement',
                'Former les commerciaux sur les arguments de vente de ces produits',
                'Créer des offres bundles avec les articles les plus vendus',
            ],
            'impact_estime': f'+{int(ca_pot/1000)}K MAD potentiel (estimation)',
            'donnees': {'nb_articles': len(articles_caches), 'ca_potentiel': round(ca_pot, 0)},
        })

    # 4. MARGE INSUFFISANTE
    if marge_pct < 20:
        recos.append({
            'id': 'reco_marge',
            'priorite': 'haute',
            'categorie': 'finance',
            'icone': 'fa-coins',
            'couleur': '#eab308',
            'titre': f'Marge brute faible : {marge_pct:.1f}% (cible > 25%)',
            'description': f'La marge brute de {marge_pct:.1f}% est en dessous de la cible de 25%. Chaque point de marge gagné représente {int(ca_total * 0.01 / 1000)}K MAD supplémentaires.',
            'actions': [
                'Auditer les articles à marge négative ou très faible (<10%)',
                'Renégocier les conditions d\'achat avec les fournisseurs principaux',
                'Réduire les remises accordées systématiquement (politique de prix ferme)',
                'Identifier les clients à qui des remises excessives sont accordées',
            ],
            'impact_estime': f'+{int(ca_total * 0.05 / 1000)}K MAD si +5pts de marge',
            'donnees': {'marge_actuelle': round(marge_pct, 1), 'marge_cible': 25, 'ca_total': round(ca_total, 0)},
        })

    # 5. RÉGIONS SOUS-PERFORMANTES
    if regions_faibles:
        noms = [r['region'] for r in regions_faibles]
        ca_manque = sum(ca_moy_region - float(r['ca'] or 0) for r in regions_faibles)
        recos.append({
            'id': 'reco_regions',
            'priorite': 'moyenne',
            'categorie': 'commercial',
            'icone': 'fa-map-marked-alt',
            'couleur': '#3b82f6',
            'titre': f'Région(s) {", ".join(noms)} sous la moyenne de {int(ca_moy_region/1000)}K MAD',
            'description': f'Les régions {", ".join(noms)} génèrent moins de 60% de la moyenne régionale. Un renforcement des ressources commerciales dans ces zones comblerait un manque de {int(ca_manque/1000)}K MAD.',
            'actions': [
                f'Affecter un commercial dédié aux régions : {", ".join(noms)}',
                'Analyser les raisons de la sous-performance (pas de présence ? concurrence forte ?)',
                'Fixer des objectifs trimestriels de rattrapage avec suivi mensuel',
                'Envisager des actions promotionnelles ciblées sur ces zones',
            ],
            'impact_estime': f'+{int(ca_manque/1000)}K MAD de CA récupérable',
            'donnees': {'regions': noms, 'ca_moyen_regional': round(ca_moy_region, 0), 'ca_manquant': round(ca_manque, 0)},
        })

    # 6. TENDANCE BAISSIÈRE
    if tendance_pct < -10:
        recos.append({
            'id': 'reco_tendance',
            'priorite': 'haute',
            'categorie': 'commercial',
            'icone': 'fa-chart-line',
            'couleur': '#ef4444',
            'titre': f'Tendance baissière détectée : {tendance_pct:+.1f}% sur la période',
            'description': f'Le CA a baissé de {abs(tendance_pct):.1f}% entre le début et la fin de la période analysée. Une action corrective immédiate est recommandée pour stopper cette tendance.',
            'actions': [
                'Organiser une réunion commerciale d\'urgence pour identifier les causes',
                'Revoir le pipeline commercial et accélérer les offres en cours',
                'Lancer une campagne de relance clients sur les comptes dormants',
                'Analyser si la baisse est saisonnière ou structurelle',
            ],
            'impact_estime': 'Stabilisation et retour à la tendance historique',
            'donnees': {'variation_pct': tendance_pct},
        })
    elif tendance_pct > 15:
        recos.append({
            'id': 'reco_tendance_positive',
            'priorite': 'faible',
            'categorie': 'commercial',
            'icone': 'fa-rocket',
            'couleur': '#22c55e',
            'titre': f'Croissance forte : {tendance_pct:+.1f}% — capitaliser sur l\'élan',
            'description': f'La tendance est fortement positive ({tendance_pct:+.1f}%). C\'est le bon moment pour investir et consolider la croissance.',
            'actions': [
                'Renforcer les stocks et la capacité de livraison pour accompagner la croissance',
                'Proposer des contrats annuels aux meilleurs clients pour sécuriser le CA',
                'Recruter ou former des commerciaux supplémentaires',
                'Explorer de nouveaux segments ou territoires pendant que la dynamique est favorable',
            ],
            'impact_estime': 'Maintien et accélération de la croissance',
            'donnees': {'variation_pct': tendance_pct},
        })

    # 7. CLIENTS INACTIFS TOTAL
    if clients_inactifs and len(clients_inactifs) != len(clients_risque):
        recos.append({
            'id': 'reco_inactifs',
            'priorite': 'faible',
            'categorie': 'client',
            'icone': 'fa-user-slash',
            'couleur': '#64748b',
            'titre': f'{len(clients_inactifs)} client(s) inactif(s) à long terme',
            'description': f'{len(clients_inactifs)} client(s) n\'ont pas commandé depuis plus de {int(rec_med * 2.5)} jours. Il faut décider de les réactiver ou de les sortir du portefeuille actif.',
            'actions': [
                'Envoyer une enquête de satisfaction pour comprendre le départ',
                'Proposer une offre de retour exceptionnelle (conditions préférentielles)',
                'Si pas de réponse sous 30j, reclasser en prospect froid',
                'Analyser si ces clients sont allés chez un concurrent identifiable',
            ],
            'impact_estime': 'Nettoyage portefeuille ou récupération ponctuelle',
            'donnees': {'nb_inactifs': len(clients_inactifs)},
        })

    # Trier : haute > moyenne > faible
    ordre = {'haute': 0, 'moyenne': 1, 'faible': 2}
    recos.sort(key=lambda r: ordre.get(r['priorite'], 3))

    # Score santé global (0-100)
    malus = 0
    if tendance_pct < -10:   malus += 25
    elif tendance_pct < 0:   malus += 10
    if marge_pct < 20:       malus += 20
    elif marge_pct < 25:     malus += 10
    if conc_top3 > 75:       malus += 15
    elif conc_top3 > 60:     malus += 8
    if clients_risque:       malus += min(20, len(clients_risque) * 5)
    if regions_faibles:      malus += len(regions_faibles) * 5
    score = max(0, 100 - malus)

    niveau = 'Excellent' if score >= 80 else 'Bon' if score >= 60 else 'Moyen' if score >= 40 else 'Critique'
    couleur_score = '#22c55e' if score >= 80 else '#3b82f6' if score >= 60 else '#eab308' if score >= 40 else '#ef4444'

    resume_parts = []
    nb_haute = sum(1 for r in recos if r['priorite'] == 'haute')
    nb_moy   = sum(1 for r in recos if r['priorite'] == 'moyenne')
    if nb_haute: resume_parts.append(f'{nb_haute} action(s) prioritaire(s)')
    if nb_moy:   resume_parts.append(f'{nb_moy} opportunité(s) à saisir')
    resume = ' · '.join(resume_parts) if resume_parts else 'Situation globalement saine.'

    return Response({
        'recommandations': recos,
        'score_sante': score,
        'niveau_sante': niveau,
        'couleur_score': couleur_score,
        'resume': resume,
        'nb_haute': nb_haute if nb_haute else 0,
        'stats_contexte': {
            'ca_total': round(ca_total, 0),
            'marge_pct': round(marge_pct, 1),
            'nb_clients': nb_clients,
            'tendance_pct': tendance_pct,
            'concentration_top3': round(conc_top3, 1),
        },
    })


# ============================================================
# API ALERTES
# ============================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_alertes(request):
    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')
    qs = _base_qs(region=region, periode=periode, config_id=config_id, date_debut=date_debut, date_fin=date_fin)

    alertes = []
    ref_date = _get_ref_date()

    monthly = list(qs
        .annotate(mois=TruncMonth('date_transaction'))
        .values('mois')
        .annotate(ca=Sum('ca_ligne'), nb=Count('id_donnee'))
        .order_by('mois')
    )

    if monthly:
        avg_ca = sum(float(m['ca'] or 0) for m in monthly) / len(monthly)
        for m in monthly[-3:]:
            ca_m = float(m['ca'] or 0)
            if ca_m < avg_ca * 0.8 and ca_m > 0:
                alertes.append({
                    'type': 'Baisse de CA',
                    'message': f"CA de {m['mois'].strftime('%b %Y')} inférieur de plus de 20% à la moyenne ({avg_ca:,.0f} MAD)",
                    'ca': round(ca_m, 2),
                    'nb_commandes': m['nb'],
                    'severite': 'warning',
                })

    # Articles sans vente le dernier mois des données (ancré sur ref_date)
    dernier_mois = ref_date.replace(day=1)
    articles_actifs = qs.filter(date_transaction__gte=dernier_mois).values('code_article').distinct().count()
    tous_articles = qs.values('code_article').distinct().count()
    inactifs = tous_articles - articles_actifs
    if inactifs > 0:
        alertes.append({
            'type': 'Articles sans vente',
            'message': f"{inactifs} article(s) sans vente en {dernier_mois.strftime('%B %Y')}",
            'ca': 0, 'nb_commandes': 0, 'severite': 'info',
        })

    return Response({'alertes': alertes})


# ============================================================
# API EXPORTS
# ============================================================

def api_export_excel(request):
    """Export Excel — accepte JWT Bearer (fetch JS) ET session Django (accès direct navigateur)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from datetime import datetime
    from rest_framework_simplejwt.authentication import JWTAuthentication

    # Auth : session Django OU JWT Bearer
    user = getattr(request, 'user', None)
    if not (user and user.is_authenticated):
        # Tentative auth JWT manuelle (pour accès direct sans session)
        try:
            jwt_auth = JWTAuthentication()
            auth_result = jwt_auth.authenticate(request)
            if auth_result:
                user, _ = auth_result
            else:
                return JsonResponse({'detail': 'Non authentifié — connectez-vous d\'abord sur /login/'}, status=401)
        except Exception:
            return JsonResponse({'detail': 'Non authentifié — connectez-vous d\'abord sur /login/'}, status=401)

    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')
    qs = _base_qs(region=region, periode=periode, config_id=config_id, date_debut=date_debut, date_fin=date_fin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Données"

    headers = ['Date', 'Code Client', 'Nom Client', 'Région', 'Code Article',
               'Nom Article', 'Catégorie', 'Quantité', 'Prix Unitaire', 'Remise', 'CA', 'Marge']
    fill = PatternFill("solid", fgColor="F97316")
    bold_white = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = bold_white

    for d in qs.order_by('date_transaction'):
        ws.append([
            str(d.date_transaction), d.code_client, d.nom_client, d.region,
            d.code_article, d.nom_article, d.categorie,
            float(d.quantite), float(d.prix_unitaire), float(d.remise),
            float(d.ca_ligne), float(d.marge_ligne),
        ])

    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    suffix = f"_{region}" if region != 'all' else ''
    suffix += f"_{periode}" if periode not in ('all', '') else ''
    date_str = datetime.now().strftime('%Y%m%d')
    filename = f"export_jumia{suffix}_{date_str}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    wb.save(response)
    return response


def api_export_csv(request):
    """Export CSV — accepte JWT Bearer (fetch JS) ET session Django (accès direct navigateur)."""
    import csv
    from datetime import datetime
    from rest_framework_simplejwt.authentication import JWTAuthentication

    # Auth : session Django OU JWT Bearer
    user = getattr(request, 'user', None)
    if not (user and user.is_authenticated):
        try:
            jwt_auth = JWTAuthentication()
            auth_result = jwt_auth.authenticate(request)
            if auth_result:
                user, _ = auth_result
            else:
                return JsonResponse({'detail': 'Non authentifié — connectez-vous d\'abord sur /login/'}, status=401)
        except Exception:
            return JsonResponse({'detail': 'Non authentifié — connectez-vous d\'abord sur /login/'}, status=401)

    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')
    qs = _base_qs(region=region, periode=periode, config_id=config_id, date_debut=date_debut, date_fin=date_fin)

    suffix = f"_{region}" if region != 'all' else ''
    suffix += f"_{periode}" if periode not in ('all', '') else ''
    date_str = datetime.now().strftime('%Y%m%d')
    filename = f"export_jumia{suffix}_{date_str}.csv"

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Date', 'Code Client', 'Nom Client', 'Région', 'Code Article',
                     'Nom Article', 'Catégorie', 'Quantité', 'Prix Unitaire', 'Remise', 'CA', 'Marge'])
    for d in qs.order_by('date_transaction'):
        writer.writerow([d.date_transaction, d.code_client, d.nom_client, d.region,
                         d.code_article, d.nom_article, d.categorie,
                         d.quantite, d.prix_unitaire, d.remise, d.ca_ligne, d.marge_ligne])
    return response


# ============================================================
# API IMPORT
# ============================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_import_excel(request):
    """
    Import flexible d'un fichier Excel (.xlsx/.xls) ou CSV.
    Détecte automatiquement les colonnes quelle que soit leur orthographe.
    Crée une ConfigurationProjet par défaut si l'utilisateur n'en a pas.

    POST multipart/form-data  :  file=<fichier>
    Réponse JSON :
    {
      "imported": 42, "errors": 0, "skipped": 1,
      "total_rows": 43, "colonnes_detectees": [...],
      "messages_erreurs": [...], "config_id": 1
    }
    """
    # ── Auth : session Django ou JWT ────────────────────────
    user = getattr(request, 'user', None)
    if not (user and user.is_authenticated):
        try:
            from rest_framework_simplejwt.authentication import JWTAuthentication
            auth_result = JWTAuthentication().authenticate(request)
            if auth_result:
                user, _ = auth_result
            else:
                return JsonResponse({'error': 'Non authentifié'}, status=401)
        except Exception:
            return JsonResponse({'error': 'Non authentifié'}, status=401)

    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

    fichier = request.FILES.get('file')
    if not fichier:
        return JsonResponse({'error': 'Aucun fichier fourni'}, status=400)

    import pandas as pd
    import unicodedata, re

    # ── Lecture du fichier ───────────────────────────────────
    nom = fichier.name.lower()
    try:
        if nom.endswith('.csv'):
            # Essai UTF-8, puis latin-1
            try:
                df = pd.read_csv(fichier, sep=None, engine='python', dtype=str)
            except Exception:
                fichier.seek(0)
                df = pd.read_csv(fichier, sep=None, engine='python', dtype=str, encoding='latin-1')
        elif nom.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(fichier, dtype=str)
        else:
            return JsonResponse({'error': 'Format non supporté. Utilisez .xlsx, .xls ou .csv'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Impossible de lire le fichier : {e}'}, status=400)

    if df.empty:
        return JsonResponse({'error': 'Le fichier est vide'}, status=400)

    # ── Normalisation des noms de colonnes ───────────────────
    def normalize(s):
        """Minuscule, supprime accents, espaces → underscore."""
        s = str(s).strip().lower()
        s = unicodedata.normalize('NFD', s)
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        s = re.sub(r'[\s\-\/\(\)%\']+', '_', s)
        s = re.sub(r'_+', '_', s).strip('_')
        return s

    col_map = {normalize(c): c for c in df.columns}

    # ── Table de correspondance normalisé → champ Django ────
    ALIASES = {
        'date_transaction':  ['date', 'date_transaction', 'date_commande', 'date_vente',
                               'transaction_date', 'order_date'],
        'code_client':       ['code_client', 'client_code', 'codeclient', 'id_client',
                               'client_id', 'num_client'],
        'nom_client':        ['nom_client', 'client', 'nom', 'raison_sociale', 'customer',
                               'name', 'entreprise'],
        'region':            ['region', 'zone', 'wilaya', 'departement', 'secteur', 'area', 'city'],
        'code_article':      ['code_article', 'article_code', 'codearticle', 'reference',
                               'ref', 'sku', 'code_produit', 'product_code'],
        'nom_article':       ['nom_article', 'article', 'produit', 'product', 'designation',
                               'libelle', 'description', 'label'],
        'categorie':         ['categorie', 'category', 'famille', 'family', 'type_produit',
                               'product_type', 'gamme'],
        'code_commercial':   ['code_commercial', 'commercial_code', 'code_vendeur',
                               'vendeur_code', 'salesperson_id', 'rep_code'],
        'nom_commercial':    ['nom_commercial', 'commercial', 'vendeur', 'salesperson',
                               'representant', 'agent'],
        'quantite':          ['quantite', 'qty', 'quantity', 'qte', 'nb', 'nombre_unites',
                               'units', 'volume'],
        'prix_unitaire':     ['prix_unitaire', 'pu', 'price', 'unit_price', 'tarif',
                               'prix', 'montant_unitaire'],
        'remise':            ['remise', 'remise_pourcentage', 'discount', 'reduction',
                               'taux_remise', 'discount_rate', 'pct_remise'],
        'ca_ligne':          ['ca_ligne', 'ca', 'montant', 'chiffre_affaires', 'amount',
                               'total', 'revenue', 'ca_ht', 'turnover'],
        'marge_ligne':       ['marge_ligne', 'marge', 'margin', 'profit', 'benefice',
                               'gross_margin'],
    }

    def find_col(field):
        """Retourne le nom de colonne original dans df pour un champ Django."""
        for alias in ALIASES.get(field, []):
            if alias in col_map:
                return col_map[alias]
        return None

    mapping = {field: find_col(field) for field in ALIASES}
    detected = [v for v in mapping.values() if v]

    # Vérifier colonnes minimales
    required = ['date_transaction', 'code_client', 'code_article']
    missing = [f for f in required if not mapping[f]]
    if missing:
        return JsonResponse({
            'error': f'Colonnes obligatoires non trouvées : {", ".join(missing)}. '
                     f'Colonnes détectées dans le fichier : {list(df.columns)}',
            'colonnes_fichier': list(df.columns),
        }, status=400)

    # ── Configuration par défaut ─────────────────────────────
    config = ConfigurationProjet.objects.filter(created_by=user).first()
    if not config:
        config = ConfigurationProjet.objects.create(
            nom_projet='Import direct',
            description='Configuration créée automatiquement lors du premier import',
            created_by=user,
        )

    # ── Import ligne par ligne ───────────────────────────────
    imported = 0
    skipped  = 0
    errors   = []

    def safe_float(val, default=0.0):
        try:
            if pd.isna(val): return default
            return float(str(val).replace(',', '.').replace(' ', ''))
        except Exception:
            return default

    def safe_str(val, default=''):
        try:
            if pd.isna(val): return default
            return str(val).strip()
        except Exception:
            return default

    def safe_date(val):
        try:
            if pd.isna(val): return None
            return pd.to_datetime(val, dayfirst=True).date()
        except Exception:
            return None

    # Champs extras (colonnes non mappées → champs_personnalises)
    mapped_cols = set(v for v in mapping.values() if v)
    extra_cols  = [c for c in df.columns if c not in mapped_cols]

    bulk_list = []
    for idx, row in df.iterrows():
        try:
            date_val = safe_date(row.get(mapping['date_transaction']))
            if date_val is None:
                skipped += 1
                continue

            code_client  = safe_str(row.get(mapping['code_client'], ''))
            code_article = safe_str(row.get(mapping['code_article'], ''))
            if not code_client or not code_article:
                skipped += 1
                continue

            quantite     = safe_float(row.get(mapping['quantite']) if mapping['quantite'] else None, 1.0)
            prix_unit    = safe_float(row.get(mapping['prix_unitaire']) if mapping['prix_unitaire'] else None, 0.0)
            remise       = safe_float(row.get(mapping['remise']) if mapping['remise'] else None, 0.0)

            # CA : lu dans le fichier ou calculé
            if mapping['ca_ligne'] and not pd.isna(row.get(mapping['ca_ligne'], float('nan'))):
                ca = safe_float(row.get(mapping['ca_ligne']))
            else:
                ca = quantite * prix_unit * (1 - remise / 100)

            # Marge : lu dans le fichier ou estimée à 25%
            if mapping['marge_ligne'] and not pd.isna(row.get(mapping['marge_ligne'], float('nan'))):
                marge = safe_float(row.get(mapping['marge_ligne']))
            else:
                marge = ca * 0.25

            extra = {}
            for ec in extra_cols:
                v = row.get(ec)
                if v is not None and not (isinstance(v, float) and pd.isna(v)):
                    extra[ec] = safe_str(v)

            bulk_list.append(DonneeBrute(
                config=config,
                date_transaction=date_val,
                code_client=code_client[:50],
                nom_client=safe_str(row.get(mapping['nom_client'], ''))[:200] if mapping['nom_client'] else '',
                region=safe_str(row.get(mapping['region'], ''))[:100] if mapping['region'] else '',
                code_article=code_article[:50],
                nom_article=safe_str(row.get(mapping['nom_article'], ''))[:200] if mapping['nom_article'] else '',
                categorie=safe_str(row.get(mapping['categorie'], ''))[:100] if mapping['categorie'] else '',
                code_commercial=safe_str(row.get(mapping['code_commercial'], ''))[:50] if mapping['code_commercial'] else '',
                nom_commercial=safe_str(row.get(mapping['nom_commercial'], ''))[:200] if mapping['nom_commercial'] else '',
                quantite=quantite,
                prix_unitaire=prix_unit,
                remise=remise,
                ca_ligne=round(ca, 2),
                marge_ligne=round(marge, 2),
                champs_personnalises=extra,
            ))
            imported += 1

        except Exception as e:
            errors.append(f'Ligne {idx + 2} : {str(e)}')
            if len(errors) >= 20:
                errors.append('… (trop d\'erreurs, arrêt du journal)')
                break

    # Bulk insert pour la performance
    if bulk_list:
        DonneeBrute.objects.bulk_create(bulk_list, batch_size=500)

    return JsonResponse({
        'imported':          imported,
        'skipped':           skipped,
        'errors':            len(errors),
        'total_rows':        len(df),
        'colonnes_detectees': detected,
        'colonnes_non_mappees': extra_cols,
        'messages_erreurs':  errors,
        'config_id':         config.id_config,
        'config_nom':        config.nom_projet,
    })


# ============================================================
# API INDICATEURS CONFIGURABLES
# ============================================================

def _indicateur_to_dict(ind):
    return {
        'id': ind.id_indicateur,
        'code': ind.code,
        'nom': ind.nom,
        'description': ind.description,
        'type_calcul': ind.type_calcul,
        'type_affichage': ind.type_affichage,
        'champ_source': ind.champ_source,
        'formule': ind.formule,
        'champ_numerateur': ind.champ_numerateur,
        'champ_denominateur': ind.champ_denominateur,
        'seuil_alerte_min': str(ind.seuil_alerte_min) if ind.seuil_alerte_min is not None else None,
        'seuil_alerte_max': str(ind.seuil_alerte_max) if ind.seuil_alerte_max is not None else None,
        'icone': ind.icone,
        'visible': ind.visible,
        'ordre_affichage': ind.ordre_affichage,
    }


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_indicateurs(request):
    if request.method == 'GET':
        inds = IndicateurPersonnalise.objects.filter(created_by=request.user)
        return Response([_indicateur_to_dict(i) for i in inds])

    data = request.data
    code = data.get('code', '').strip()
    nom = data.get('nom', '').strip()
    if not code or not nom:
        return Response({'error': 'Code et nom sont obligatoires'}, status=400)
    if IndicateurPersonnalise.objects.filter(code=code).exists():
        return Response({'error': f'Le code "{code}" existe déjà'}, status=400)

    ind = IndicateurPersonnalise.objects.create(
        code=code, nom=nom,
        description=data.get('description', ''),
        type_calcul=data.get('type_calcul', 'somme'),
        type_affichage=data.get('type_affichage', 'montant'),
        champ_source=data.get('champ_source', ''),
        formule=data.get('formule', ''),
        champ_numerateur=data.get('champ_numerateur', ''),
        champ_denominateur=data.get('champ_denominateur', ''),
        seuil_alerte_min=data.get('seuil_alerte_min') or None,
        seuil_alerte_max=data.get('seuil_alerte_max') or None,
        icone=data.get('icone', 'fa-chart-line'),
        created_by=request.user,
    )
    return Response(_indicateur_to_dict(ind), status=201)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def api_indicateur_detail(request, pk):
    try:
        ind = IndicateurPersonnalise.objects.get(id_indicateur=pk, created_by=request.user)
    except IndicateurPersonnalise.DoesNotExist:
        return Response({'error': 'Indicateur non trouvé'}, status=404)

    if request.method == 'GET':
        return Response(_indicateur_to_dict(ind))

    if request.method == 'PUT':
        d = request.data
        for field in ['nom', 'description', 'type_calcul', 'type_affichage',
                      'champ_source', 'formule', 'champ_numerateur', 'champ_denominateur',
                      'icone', 'visible', 'ordre_affichage']:
            if field in d:
                setattr(ind, field, d[field])
        ind.seuil_alerte_min = d.get('seuil_alerte_min') or None
        ind.seuil_alerte_max = d.get('seuil_alerte_max') or None
        ind.save()
        return Response(_indicateur_to_dict(ind))

    ind.delete()
    return Response(status=204)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_calculer_kpi(request, pk):
    try:
        ind = IndicateurPersonnalise.objects.get(id_indicateur=pk, created_by=request.user)
    except IndicateurPersonnalise.DoesNotExist:
        return Response({'error': 'Indicateur non trouvé'}, status=404)

    engine = KPIEngine()
    result = engine.calculer_kpi(pk, filtres=request.data)
    return Response(result)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_kpi_par_dimension(request, pk):
    dimension = request.GET.get('dimension', 'region')
    engine = KPIEngine()
    result = engine.calculer_par_dimension(pk, dimension)
    return Response(result)


# ============================================================
# API KPIs PERSONNALISÉS (dashboard)
# ============================================================

COLONNES_NUMERIQUES = [
    {'nom': 'ca_ligne',      'label': 'CA Ligne',       'type': 'montant'},
    {'nom': 'marge_ligne',   'label': 'Marge Ligne',    'type': 'montant'},
    {'nom': 'quantite',      'label': 'Quantité',       'type': 'nombre'},
    {'nom': 'prix_unitaire', 'label': 'Prix Unitaire',  'type': 'montant'},
    {'nom': 'remise',        'label': 'Remise (%)',      'type': 'pourcentage'},
]

COLONNES_COMPTE = [
    {'nom': 'code_client',     'label': 'Clients (distinct)',     'type': 'texte'},
    {'nom': 'code_article',    'label': 'Articles (distinct)',    'type': 'texte'},
    {'nom': 'code_commercial', 'label': 'Commerciaux (distinct)', 'type': 'texte'},
    {'nom': 'region',          'label': 'Régions (distinct)',     'type': 'texte'},
    {'nom': 'categorie',       'label': 'Catégories (distinct)',  'type': 'texte'},
]


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_colonnes_disponibles(request):
    """Retourne les colonnes disponibles pour configurer un KPI."""
    config_id = request.GET.get('config_id')
    qs = DonneeBrute.objects.all()
    if config_id:
        qs = qs.filter(config_id=config_id)

    # Scanner les clés de champs_personnalises
    perso_keys = set()
    for d in qs.values_list('champs_personnalises', flat=True)[:200]:
        if d:
            perso_keys.update(d.keys())

    champs_perso = [{'nom': k, 'label': k, 'type': 'texte'} for k in sorted(perso_keys)]

    return Response({
        'colonnes_numeriques': COLONNES_NUMERIQUES,
        'colonnes_compte': COLONNES_COMPTE,
        'champs_personnalises': champs_perso,
    })


def _eval_formule_kpi(formule, qs):
    ca = float(qs.aggregate(r=Sum('ca_ligne'))['r'] or 0)
    marge = float(qs.aggregate(r=Sum('marge_ligne'))['r'] or 0)
    nb = qs.aggregate(r=Count('id_donnee'))['r'] or 0
    qte = float(qs.aggregate(r=Sum('quantite'))['r'] or 0)
    ctx = {
        'ca': ca, 'marge': marge, 'nb_commandes': float(nb),
        'quantite': qte, 'panier_moyen': ca / nb if nb else 0,
        'round': round, 'abs': abs, 'max': max, 'min': min,
    }
    try:
        return float(eval(formule, {"__builtins__": {}}, ctx))
    except Exception:
        return 0


def _calculer_indicateur(ind, qs):
    try:
        tc = ind.type_calcul
        champ = ind.champ_source
        if tc == 'somme':
            return float(qs.aggregate(r=Sum(champ))['r'] or 0)
        elif tc == 'moyenne':
            return float(qs.aggregate(r=Avg(champ))['r'] or 0)
        elif tc == 'compte':
            return qs.values(champ).distinct().count()
        elif tc == 'min':
            return float(qs.aggregate(r=Min(champ))['r'] or 0)
        elif tc == 'max':
            return float(qs.aggregate(r=Max(champ))['r'] or 0)
        elif tc == 'ratio':
            num = float(qs.aggregate(r=Sum(ind.champ_numerateur))['r'] or 0)
            den = float(qs.aggregate(r=Sum(ind.champ_denominateur))['r'] or 1)
            return round(num / den, 4) if den else 0
        elif tc == 'formule':
            return _eval_formule_kpi(ind.formule, qs)
        return 0
    except Exception:
        return None


def _check_alerte_kpi(ind, valeur):
    if valeur is None:
        return None
    if ind.seuil_alerte_min is not None and valeur < float(ind.seuil_alerte_min):
        return 'bas'
    if ind.seuil_alerte_max is not None and valeur > float(ind.seuil_alerte_max):
        return 'haut'
    return None


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_kpis_personnalises(request):
    """Calcule tous les KPIs personnalisés de l'utilisateur avec les filtres du dashboard."""
    region = request.GET.get('region', 'all')
    periode = request.GET.get('periode', 'all')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    config_id = request.GET.get('config_id')

    qs = _base_qs(region=region, periode=periode, config_id=config_id,
                  date_debut=date_debut, date_fin=date_fin)
    inds = (IndicateurPersonnalise.objects
            .filter(created_by=request.user, visible=True)
            .order_by('ordre_affichage', 'nom'))

    results = []
    for ind in inds:
        valeur = _calculer_indicateur(ind, qs)
        results.append({
            'id': ind.id_indicateur,
            'nom': ind.nom,
            'description': ind.description,
            'type_affichage': ind.type_affichage,
            'icone': ind.icone,
            'valeur': valeur,
            'alerte': _check_alerte_kpi(ind, valeur),
            'seuil_min': float(ind.seuil_alerte_min) if ind.seuil_alerte_min is not None else None,
            'seuil_max': float(ind.seuil_alerte_max) if ind.seuil_alerte_max is not None else None,
        })

    return Response(results)


# ============================================================
# API CONFIGURATIONS PROJET
# ============================================================

def _config_to_dict(c):
    return {
        'id': c.id_config,
        'nom': c.nom_projet,
        'description': c.description,
        'theme': c.theme_couleur,
        'created_at': c.created_at.isoformat(),
    }


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_configurations(request):
    if request.method == 'GET':
        return Response([_config_to_dict(c) for c in ConfigurationProjet.objects.filter(created_by=request.user)])

    config = ConfigurationProjet.objects.create(
        nom_projet=request.data.get('nom', 'Mon Dashboard'),
        description=request.data.get('description', ''),
        theme_couleur=request.data.get('theme', '#f97316'),
        created_by=request.user,
    )

    # Créer des indicateurs et widgets par défaut
    _creer_widgets_par_defaut(request.user, config)

    return Response(_config_to_dict(config), status=201)


def _creer_widgets_par_defaut(user, config):
    """Crée des indicateurs et widgets KPI par défaut pour une nouvelle configuration."""
    indicateurs_defaut = [
        {
            'nom': 'Chiffre d\'Affaires Total',
            'code': f'CA_TOTAL_{config.id_config}',
            'type_calcul': 'somme',
            'champ_source': 'ca_ligne',
            'type_affichage': 'montant',
            'icone': 'fa-chart-line',
            'description': 'Somme du chiffre d\'affaires total',
        },
        {
            'nom': 'Marge Totale',
            'code': f'MARGE_{config.id_config}',
            'type_calcul': 'somme',
            'champ_source': 'marge_ligne',
            'type_affichage': 'montant',
            'icone': 'fa-coins',
            'description': 'Marge brute totale',
        },
        {
            'nom': 'Nombre de Commandes',
            'code': f'NB_CMD_{config.id_config}',
            'type_calcul': 'compte',
            'champ_source': 'id_donnee',
            'type_affichage': 'nombre',
            'icone': 'fa-shopping-cart',
            'description': 'Nombre total de transactions',
        },
    ]

    types_widget = ['kpi_card', 'bar_chart', 'line_chart']
    for i, ind_data in enumerate(indicateurs_defaut):
        try:
            # Ne créer que si le code n'existe pas encore
            ind, created = IndicateurPersonnalise.objects.get_or_create(
                code=ind_data['code'],
                created_by=user,
                defaults={
                    'nom': ind_data['nom'],
                    'type_calcul': ind_data['type_calcul'],
                    'champ_source': ind_data['champ_source'],
                    'type_affichage': ind_data['type_affichage'],
                    'icone': ind_data['icone'],
                    'description': ind_data['description'],
                    'table_source': 'donnee_brute',
                }
            )
            if created:
                WidgetDashboard.objects.create(
                    nom=ind_data['nom'],
                    type_widget=types_widget[i],
                    indicateur=ind,
                    largeur=4,
                    hauteur=3,
                    created_by=user,
                )
        except Exception:
            pass  # Ignorer les erreurs de création de widgets par défaut


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def api_configuration_detail(request, pk):
    try:
        config = ConfigurationProjet.objects.get(id_config=pk, created_by=request.user)
    except ConfigurationProjet.DoesNotExist:
        return Response({'error': 'Configuration non trouvée'}, status=404)

    if request.method == 'GET':
        return Response(_config_to_dict(config))
    if request.method == 'PUT':
        config.nom_projet = request.data.get('nom', config.nom_projet)
        config.description = request.data.get('description', config.description)
        config.theme_couleur = request.data.get('theme', config.theme_couleur)
        config.save()
        return Response(_config_to_dict(config))
    config.delete()
    return Response(status=204)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_generer_canevas(request, pk):
    try:
        config = ConfigurationProjet.objects.get(id_config=pk, created_by=request.user)
    except ConfigurationProjet.DoesNotExist:
        return Response({'error': 'Configuration non trouvée'}, status=404)

    engine = KPIEngine(config_id=pk)
    output = engine.generer_canevas_excel(pk)
    if not output:
        return Response({'error': 'Erreur lors de la génération'}, status=500)

    nom = config.nom_projet.replace(' ', '_').replace('/', '-')
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="canevas_{nom}.xlsx"'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_importer_canevas(request, pk):
    try:
        ConfigurationProjet.objects.get(id_config=pk, created_by=request.user)
    except ConfigurationProjet.DoesNotExist:
        return Response({'error': 'Configuration non trouvée'}, status=404)

    fichier = request.FILES.get('file')
    if not fichier:
        return Response({'error': 'Aucun fichier fourni'}, status=400)

    engine = KPIEngine(config_id=pk)
    result = engine.importer_canevas(pk, fichier)

    if 'erreur' in result:
        return Response({'error': result['erreur']}, status=400)

    return Response({
        'imported': result.get('imported', 0),
        'errors': result.get('errors', []),
        'total_rows': result.get('total_rows', 0),
    })


# ============================================================
# API CANEVAS LIBRE (génération sans configuration préalable)
# ============================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_generer_canevas_libre(request):
    """
    Génère un fichier Excel canevas personnalisé à partir des colonnes
    choisies par l'utilisateur — sans nécessiter de ConfigurationProjet.
    """
    import pandas as pd
    from io import BytesIO

    body = request.data
    if not isinstance(body, dict):
        return Response({'detail': 'JSON invalide'}, status=400)

    colonnes = body.get('colonnes', [])
    nom_fichier = body.get('nom_fichier', 'canevas_personnalise').replace(' ', '_')

    if not colonnes:
        return Response({'detail': 'Au moins une colonne est requise'}, status=400)

    col_noms = [c['nom'] for c in colonnes]

    # ── Lignes d'exemple (2 lignes) ──────────────────────────
    EXEMPLES_PAR_TYPE = {
        'date':        ['2024-01-15', '2024-01-16'],
        'texte':       ['Exemple 1', 'Exemple 2'],
        'nombre':      [10, 5],
        'montant':     [15000, 8500],
        'pourcentage': [10, 5],
    }
    # Correspondances colonnes standard → exemples spécifiques
    EXEMPLES_SPECIAUX = {
        'Date':             ['2024-01-15', '2024-01-16'],
        'Code Client':      ['CLT001', 'CLT002'],
        'Nom Client':       ['Entreprise Alpha', 'Société Beta'],
        'Région':           ['Alger', 'Oran'],
        'Code Article':     ['ART001', 'ART002'],
        'Nom Article':      ['Produit A', 'Produit B'],
        'Catégorie':        ['Électronique', 'Maison'],
        'Code Commercial':  ['COM001', 'COM002'],
        'Nom Commercial':   ['Ahmed Benali', 'Sara Meziane'],
        'Quantité':         [5, 3],
        'Prix Unitaire':    [10000, 15000],
        'Remise (%)':       [10, 5],
        'Marge Ligne':      [2500, 1800],
    }

    rows = []
    for i in range(2):
        row = {}
        for c in colonnes:
            nom = c['nom']
            typ = c.get('type', 'texte')
            if nom in EXEMPLES_SPECIAUX:
                row[nom] = EXEMPLES_SPECIAUX[nom][i]
            else:
                vals = EXEMPLES_PAR_TYPE.get(typ, ['Exemple 1', 'Exemple 2'])
                row[nom] = vals[i]
        rows.append(row)

    df = pd.DataFrame(rows, columns=col_noms)

    # ── Feuille Instructions ──────────────────────────────────
    TYPE_LABELS = {
        'date': 'Date (YYYY-MM-DD)', 'texte': 'Texte libre',
        'nombre': 'Nombre entier', 'montant': 'Montant (MAD)',
        'pourcentage': 'Pourcentage (%)',
    }
    instructions = pd.DataFrame({
        'Colonne':     [c['nom'] for c in colonnes],
        'Type':        [TYPE_LABELS.get(c.get('type', 'texte'), 'Texte') for c in colonnes],
        'Obligatoire': ['Oui' if c.get('obligatoire') else 'Non' for c in colonnes],
        'Description': [c.get('description', '') for c in colonnes],
    })

    # ── Export Excel ──────────────────────────────────────────
    output = BytesIO()
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Données à saisir', index=False)
            instructions.to_excel(writer, sheet_name='Instructions', index=False)

            # ── Mise en forme feuille Données ──
            ws = writer.sheets['Données à saisir']

            ORANGE  = 'F97316'
            ORANGE_L = 'FFF3E0'
            GRAY    = 'F0F4FF'
            RED     = 'EF4444'
            RED_L   = 'FFF5F5'

            header_fill_oblig = PatternFill('solid', fgColor=ORANGE)
            header_fill_opt   = PatternFill('solid', fgColor='1E293B')
            data_fill_oblig   = PatternFill('solid', fgColor=ORANGE_L)
            data_fill_opt     = PatternFill('solid', fgColor='FAFAFA')

            # Colonnes obligatoires marquées en orange
            obligatoires = {c['nom'] for c in colonnes if c.get('obligatoire')}

            thin = Side(style='thin', color='DDDDDD')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for ci, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
                cell = col[0]
                is_oblig = cell.value in obligatoires
                cell.fill      = header_fill_oblig if is_oblig else header_fill_opt
                cell.font      = Font(bold=True, color='FFFFFF', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border    = border

            # Lignes d'exemple
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    col_name = ws.cell(1, cell.column).value
                    is_oblig = col_name in obligatoires
                    cell.fill      = data_fill_oblig if is_oblig else data_fill_opt
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border    = border

            # Hauteur header + ajustement largeur colonnes
            ws.row_dimensions[1].height = 30
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

            # Figer la 1ère ligne
            ws.freeze_panes = 'A2'

            # ── Mise en forme feuille Instructions ──
            ws2 = writer.sheets['Instructions']
            instr_fill = PatternFill('solid', fgColor='1E293B')
            for cell in ws2[1]:
                cell.fill = instr_fill
                cell.font = Font(bold=True, color='FFFFFF')
            for col in ws2.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    except ImportError:
        # Fallback sans mise en forme si openpyxl absent
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Données à saisir', index=False)
            instructions.to_excel(writer, sheet_name='Instructions', index=False)

    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{nom_fichier}.xlsx"'
    )
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    return response


# ============================================================
# API WIDGETS
# ============================================================

def _widget_to_dict(w):
    return {
        'id': w.id_widget,
        'nom': w.nom,
        'type': w.type_widget,
        'indicateur': w.indicateur.nom,
        'indicateur_id': w.indicateur.id_indicateur,
        'taille': {'w': w.largeur, 'h': w.hauteur},
    }


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_widgets(request):
    if request.method == 'GET':
        return Response([_widget_to_dict(w) for w in WidgetDashboard.objects.filter(created_by=request.user)])

    data = request.data
    try:
        ind = IndicateurPersonnalise.objects.get(id_indicateur=data.get('indicateur_id'), created_by=request.user)
    except IndicateurPersonnalise.DoesNotExist:
        return Response({'error': 'Indicateur non trouvé'}, status=404)

    widget = WidgetDashboard.objects.create(
        nom=data.get('nom', ''), type_widget=data.get('type_widget', 'kpi_card'),
        indicateur=ind, largeur=data.get('largeur', 6), hauteur=data.get('hauteur', 4),
        created_by=request.user,
    )
    return Response(_widget_to_dict(widget), status=201)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_dashboard_dynamique(request):
    widgets = WidgetDashboard.objects.filter(created_by=request.user)
    return Response({'widgets': [_widget_to_dict(w) for w in widgets]})


# ============================================================
# SETUP RAILWAY
# ============================================================

def setup_railway(request):
    results = []
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
        results.append("✅ Base de données OK")

        call_command('migrate', '--noinput')
        results.append("✅ Migrations appliquées")

        User = get_user_model()
        admin, created = User.objects.get_or_create(username='admin', defaults={'is_superuser': True, 'is_staff': True, 'email': ''})
        admin.set_password('admin123')
        admin.is_superuser = True
        admin.is_staff = True
        admin.is_active = True
        admin.save()
        action = "créé" if created else "réinitialisé"
        results.append(f"✅ Admin {action} → login: admin / admin123")

        call_command('collectstatic', '--noinput', '--clear')
        results.append("✅ Fichiers statiques collectés")

        return JsonResponse({'status': 'success', 'details': results,
                             'credentials': {'username': 'admin', 'password': 'admin123'}})
    except Exception as e:
        results.append(f"Erreur: {e}")
        return JsonResponse({'status': 'error', 'details': results})


# ============================================================
# IMPORT ETAT DATA (endpoint production)
# ============================================================

def import_etat_data(request):
    """
    Endpoint pour importer les donnees ETAT.xlsx en production.
    Protege par token secret.
    """
    import json
    from pathlib import Path
    from django.db import transaction as db_transaction
    from django.db.models import Sum, Count, Min, Max

    # Verifier le token
    token = request.GET.get('token', '')
    expected = os.environ.get('IMPORT_SECRET', 'import-etat-2025')
    if token != expected:
        return JsonResponse({'status': 'error', 'message': 'Token invalide'}, status=403)

    try:
        # Chemin du fichier fixture
        fixture_path = Path(__file__).resolve().parent.parent / 'analytics' / 'fixtures' / 'etat_data_raw.json'

        if not fixture_path.exists():
            return JsonResponse({'status': 'error', 'message': f'Fichier introuvable: {fixture_path}'})

        with open(fixture_path, 'r', encoding='utf-8') as f:
            export = json.load(f)

        donnees = export.get('donnees', [])

        # Recuperer l'admin
        User = get_user_model()
        admin = User.objects.filter(is_superuser=True).first()

        from analytics.models import ConfigurationProjet, DonneeBrute

        config, created = ConfigurationProjet.objects.get_or_create(
            nom_projet='ETAT - Donnees importees',
            defaults={
                'description': 'Donnees depuis ETAT.xlsx (2021-2025)',
                'colonnes_canevas': [],
                'created_by': admin,
            }
        )

        # Creer les widgets par defaut si aucun widget pour cet admin
        if admin:
            if not WidgetDashboard.objects.filter(created_by=admin).exists():
                try:
                    _creer_widgets_par_defaut(admin, config)
                except Exception:
                    pass

        # Supprimer existants
        DonneeBrute.objects.filter(config=config).delete()

        # Importer
        batch = []
        errors = []
        for d in donnees:
            try:
                batch.append(DonneeBrute(
                    config=config,
                    date_transaction=d['date_transaction'],
                    code_client=d.get('code_client', ''),
                    nom_client=d.get('nom_client', ''),
                    region=d.get('region', ''),
                    code_article=d.get('code_article', ''),
                    nom_article=d.get('nom_article', ''),
                    categorie=d.get('categorie', ''),
                    code_commercial=d.get('code_commercial', ''),
                    nom_commercial=d.get('nom_commercial', ''),
                    quantite=float(d.get('quantite', 0)),
                    prix_unitaire=float(d.get('prix_unitaire', 0)),
                    remise=float(d.get('remise', 0)),
                    ca_ligne=float(d.get('ca_ligne', 0)),
                    marge_ligne=float(d.get('marge_ligne', 0)),
                    champs_personnalises=d.get('champs_personnalises', {}),
                ))
            except Exception as e:
                errors.append(str(e))

        with db_transaction.atomic():
            DonneeBrute.objects.bulk_create(batch, batch_size=100)

        # Statistiques
        s = DonneeBrute.objects.filter(config=config).aggregate(
            ca=Sum('ca_ligne'), nb=Count('id_donnee'),
            clt=Count('code_client', distinct=True),
            dmin=Min('date_transaction'), dmax=Max('date_transaction')
        )

        return JsonResponse({
            'status': 'success',
            'config_id': config.id_config,
            'imported': len(batch),
            'errors': len(errors),
            'stats': {
                'ca_total': float(s['ca'] or 0),
                'transactions': s['nb'],
                'clients': s['clt'],
                'periode': f"{s['dmin']} -> {s['dmax']}",
            }
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
